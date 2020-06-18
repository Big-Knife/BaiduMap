# coding=utf-8
from importlib import reload

from openpyxl import Workbook
from openpyxl import load_workbook
# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
import re, pandas as pd
import time
from bs4 import BeautifulSoup
import sys

chrome_driver = 'C:\\Users\\ITER\\Desktop\\BaiduMap\\chromedriver.exe'
browser = webdriver.Chrome(executable_path=chrome_driver)

reload(sys)



def coordinate():
    # 创建浏览器驱动对象
    driver = webdriver.Chrome()
    driver.get('http://api.map.baidu.com/lbsapi/getpoint/index.html')
    # 显式等待，设置timeout
    wait = WebDriverWait(driver, 1)  # 等待的最大时间
    # 判断输入框是否加载
    input = wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, '#localvalue')))
    # 判断搜索按钮是否加载
    submit = wait.until(
        EC.element_to_be_clickable(
            (By.CSS_SELECTOR, '#localsearch')))
    # 输入搜索词，点击搜索按钮

    # 有时候我们希望读取到公式计算出来的结果，可以使用load_workbook()中的data_only属性
    wb = load_workbook('r C:\\Users\\ITER\\Desktop\\Workbook.xlsx', data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows():
        rows.append(row)

    print
    u"行高：", ws.max_row
    print
    u"列宽：", ws.max_column
    for i in range(1, ws.max_row):  # row
        # print  rows[i][0], rows[i][0].value, type(rows[i][0].value)
        print
        rows[i][0], rows[i][0].value,
        # coordinate(rows[i][0].value)
        input.clear()
        input.send_keys(rows[i][0].value)  # u'浙江工业大学'
        submit.click()
        time.sleep(1)
        try:
            # 等待坐标
            wait.until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, '#no_0')))
        except TimeoutException:
            print('百度地图查不到地址')
            continue
        # 获取网页文本，提取经纬度
        source = driver.page_source
        soup = BeautifulSoup(source, 'lxml')  #
        i = 0
        for li in soup.select('ul.local_s > li'):
            print
            li.get_text()
            i += 1
            if i > 0:
                break
    # 关闭浏览器驱动
    driver.close()


coordinate()